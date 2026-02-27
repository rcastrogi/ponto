import os
import io
from datetime import datetime, date, timedelta
from functools import wraps
from zoneinfo import ZoneInfo

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_file, send_from_directory
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from models import get_db, init_db

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'piticas-ponto-secret-key-2026')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=10)

# Upload config
DATA_DIR = os.environ.get('DATA_DIR', os.path.dirname(os.path.abspath(__file__)))
UPLOAD_FOLDER = os.path.join(DATA_DIR, 'uploads', 'atestados')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Timezone Brasil
BR_TZ = ZoneInfo('America/Sao_Paulo')

def agora():
    """Retorna datetime atual no fuso de São Paulo."""
    return datetime.now(BR_TZ)

def hoje():
    """Retorna date atual no fuso de São Paulo."""
    return datetime.now(BR_TZ).date()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Faça login para acessar o sistema.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def gestor_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Faça login para acessar o sistema.', 'warning')
            return redirect(url_for('login'))
        if not session.get('is_gestor'):
            flash('Acesso restrito a gestores.', 'danger')
            return redirect(url_for('meu_ponto'))
        return f(*args, **kwargs)
    return decorated


def is_feriado(d, db=None):
    """Verifica se uma data é feriado."""
    close_db = False
    if db is None:
        db = get_db()
        close_db = True
    result = db.execute(
        'SELECT descricao FROM feriados WHERE data = ?', (d.isoformat(),)
    ).fetchone()
    if close_db:
        db.close()
    return result['descricao'] if result else None


def tipo_dia(d, db=None):
    """Retorna 'especial' para domingos/feriados, 'normal' para os demais."""
    if d.weekday() == 6:  # domingo
        return 'especial'
    if is_feriado(d, db):
        return 'especial'
    return 'normal'


def carga_esperada_dia(d, colaborador, db=None):
    """Retorna carga horária esperada para o dia: 8h normal, 6h dom/feriado."""
    if tipo_dia(d, db) == 'especial':
        return colaborador['horas_dia_especial'] or 6.0
    return colaborador['horas_dia_normal'] or 8.0


def calcular_horas(entrada, saida_almoco, retorno_almoco, saida):
    """Calcula total de horas trabalhadas considerando almoço."""
    if not entrada or not saida:
        if entrada and not saida:
            fmt = '%H:%M'
            h_entrada = datetime.strptime(entrada, fmt)
            agora_dt = agora()
            h_agora = datetime.strptime(agora_dt.strftime('%H:%M'), fmt)

            total_minutos = (h_agora - h_entrada).total_seconds() / 60

            if saida_almoco and retorno_almoco:
                h_saida_almoco = datetime.strptime(saida_almoco, fmt)
                h_retorno_almoco = datetime.strptime(retorno_almoco, fmt)
                almoco_minutos = (h_retorno_almoco - h_saida_almoco).total_seconds() / 60
                total_minutos -= almoco_minutos
            elif saida_almoco and not retorno_almoco:
                h_saida_almoco = datetime.strptime(saida_almoco, fmt)
                almoco_minutos = (h_agora - h_saida_almoco).total_seconds() / 60
                total_minutos -= almoco_minutos

            return round(max(0, total_minutos / 60), 2)
        return 0.0

    fmt = '%H:%M'
    try:
        h_entrada = datetime.strptime(entrada, fmt)
        h_saida = datetime.strptime(saida, fmt)
        total_minutos = (h_saida - h_entrada).total_seconds() / 60

        if saida_almoco and retorno_almoco:
            h_saida_almoco = datetime.strptime(saida_almoco, fmt)
            h_retorno_almoco = datetime.strptime(retorno_almoco, fmt)
            almoco_minutos = (h_retorno_almoco - h_saida_almoco).total_seconds() / 60
            total_minutos -= almoco_minutos

        return round(max(0, total_minutos / 60), 2)
    except (ValueError, TypeError):
        return 0.0


def calcular_horas_extras_semana(horas_semana, max_horas_semana=40.0):
    """Calcula horas extras na semana (acima do limite)."""
    if horas_semana > max_horas_semana:
        return round(horas_semana - max_horas_semana, 2)
    return 0.0


def get_semana_inicio_fim(d):
    """Retorna segunda e domingo da semana de uma data."""
    inicio = d - timedelta(days=d.weekday())
    fim = inicio + timedelta(days=6)
    return inicio, fim


def get_mes_inicio_fim(d):
    """Retorna primeiro e último dia do mês."""
    inicio = d.replace(day=1)
    if d.month == 12:
        fim = d.replace(year=d.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        fim = d.replace(month=d.month + 1, day=1) - timedelta(days=1)
    return inicio, fim


def calcular_resumo_semana(registros_semana, max_horas_semana=40.0):
    """Calcula resumo da semana: horas normais, extras, folgas usadas."""
    horas_total = sum(r['horas_trabalhadas'] for r in registros_semana)
    dias_trabalhados = len(registros_semana)
    horas_extras = calcular_horas_extras_semana(horas_total, max_horas_semana)
    horas_normais = min(horas_total, max_horas_semana)
    return {
        'horas_total': round(horas_total, 2),
        'horas_normais': round(horas_normais, 2),
        'horas_extras': horas_extras,
        'dias_trabalhados': dias_trabalhados,
    }


def calcular_horas_justificadas(colab_id, data_inicio, data_fim, colaborador, db):
    """Calcula total de horas justificadas (aprovadas) em um período.
    Para cada dia coberto por uma justificativa aprovada, soma a carga horária esperada."""
    justificativas = db.execute(
        '''SELECT * FROM justificativas
           WHERE colaborador_id = ? AND status = 'aprovado'
           AND (data_inicio <= ? AND data_fim >= ?)''',
        (colab_id, data_fim.isoformat(), data_inicio.isoformat())
    ).fetchall()

    horas = 0.0
    dias_justificados = set()
    for j in justificativas:
        j_inicio = max(date.fromisoformat(j['data_inicio']), data_inicio)
        j_fim = min(date.fromisoformat(j['data_fim']), data_fim)
        d = j_inicio
        while d <= j_fim:
            if d.isoformat() not in dias_justificados:
                dias_justificados.add(d.isoformat())
                horas += carga_esperada_dia(d, colaborador, db)
            d += timedelta(days=1)
    return round(horas, 2), len(dias_justificados)


@app.context_processor
def inject_globals():
    """Inject global variables into templates."""
    return {
        'now': agora(),
        'today': hoje(),
    }


# ---------------------------------------------------------------------------
# Auth Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('is_gestor'):
            return redirect(url_for('dashboard'))
        return redirect(url_for('meu_ponto'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')

        db = get_db()
        user = db.execute(
            'SELECT * FROM colaboradores WHERE email = ? AND ativo = 1', (email,)
        ).fetchone()
        db.close()

        if not user:
            flash('E-mail ou senha inválidos.', 'danger')
            return render_template('login.html')

        # Primeiro acesso: redirecionar para criar senha
        if user['primeiro_acesso']:
            session['primeiro_acesso_user_id'] = user['id']
            session['primeiro_acesso_user_nome'] = user['nome']
            return redirect(url_for('criar_senha'))

        if check_password_hash(user['senha'], senha):
            session.permanent = True
            session['user_id'] = user['id']
            session['user_nome'] = user['nome']
            session['is_gestor'] = bool(user['is_gestor'])

            if user['is_gestor']:
                return redirect(url_for('dashboard'))
            return redirect(url_for('meu_ponto'))

        flash('E-mail ou senha inválidos.', 'danger')
    return render_template('login.html')


@app.route('/criar-senha', methods=['GET', 'POST'])
def criar_senha():
    user_id = session.get('primeiro_acesso_user_id')
    user_nome = session.get('primeiro_acesso_user_nome', '')

    if not user_id:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nova_senha = request.form.get('nova_senha', '')
        confirmar = request.form.get('confirmar_senha', '')

        if nova_senha != confirmar:
            flash('As senhas não conferem.', 'danger')
            return render_template('criar_senha.html', nome=user_nome)

        if len(nova_senha) < 4:
            flash('A senha deve ter pelo menos 4 caracteres.', 'danger')
            return render_template('criar_senha.html', nome=user_nome)

        db = get_db()
        db.execute(
            'UPDATE colaboradores SET senha = ?, primeiro_acesso = 0 WHERE id = ?',
            (generate_password_hash(nova_senha), user_id)
        )
        db.commit()

        user = db.execute('SELECT * FROM colaboradores WHERE id = ?', (user_id,)).fetchone()
        db.close()

        # Limpar dados temporários e fazer login
        session.pop('primeiro_acesso_user_id', None)
        session.pop('primeiro_acesso_user_nome', None)

        session.permanent = True
        session['user_id'] = user['id']
        session['user_nome'] = user['nome']
        session['is_gestor'] = bool(user['is_gestor'])

        flash('Senha criada com sucesso! Bem-vindo(a)!', 'success')
        if user['is_gestor']:
            return redirect(url_for('dashboard'))
        return redirect(url_for('meu_ponto'))

    return render_template('criar_senha.html', nome=user_nome)


@app.route('/logout')
def logout():
    session.clear()
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# Punch Clock (Batida de Ponto)
# ---------------------------------------------------------------------------

@app.route('/meu-ponto')
@login_required
def meu_ponto():
    db = get_db()
    hoje_dt = hoje()
    hoje_iso = hoje_dt.isoformat()
    user_id = session['user_id']

    # Dados do colaborador
    colaborador = db.execute(
        'SELECT * FROM colaboradores WHERE id = ?', (user_id,)
    ).fetchone()

    # Registro de hoje
    registro_hoje = db.execute(
        'SELECT * FROM registros_ponto WHERE colaborador_id = ? AND data = ?',
        (user_id, hoje_iso)
    ).fetchone()

    # Registros da semana
    inicio_sem, fim_sem = get_semana_inicio_fim(hoje_dt)
    registros_semana = db.execute(
        '''SELECT * FROM registros_ponto
           WHERE colaborador_id = ? AND data BETWEEN ? AND ?
           ORDER BY data''',
        (user_id, inicio_sem.isoformat(), fim_sem.isoformat())
    ).fetchall()

    # Registros do mês
    inicio_mes, fim_mes = get_mes_inicio_fim(hoje_dt)
    registros_mes = db.execute(
        '''SELECT * FROM registros_ponto
           WHERE colaborador_id = ? AND data BETWEEN ? AND ?
           ORDER BY data''',
        (user_id, inicio_mes.isoformat(), fim_mes.isoformat())
    ).fetchall()

    # Calcular totais
    max_horas = colaborador['max_horas_semana'] or 40.0
    horas_semana_trabalhadas = sum(r['horas_trabalhadas'] for r in registros_semana)
    horas_mes_trabalhadas = sum(r['horas_trabalhadas'] for r in registros_mes)

    # Horas justificadas (aprovadas)
    horas_just_semana, _ = calcular_horas_justificadas(
        user_id, inicio_sem, fim_sem, colaborador, db)
    horas_just_mes, _ = calcular_horas_justificadas(
        user_id, inicio_mes, fim_mes, colaborador, db)

    horas_semana = horas_semana_trabalhadas + horas_just_semana
    horas_mes = horas_mes_trabalhadas + horas_just_mes
    resumo_sem = calcular_resumo_semana(registros_semana, max_horas)
    # Adicionar horas justificadas ao resumo
    resumo_sem['horas_justificadas'] = horas_just_semana
    resumo_sem['horas_total'] = round(resumo_sem['horas_total'] + horas_just_semana, 2)

    # Horas extras do mês (somar por semana incluindo justificadas)
    horas_extras_mes = 0.0
    semanas_no_mes = {}
    for r in registros_mes:
        d = date.fromisoformat(r['data'])
        sem_inicio, _ = get_semana_inicio_fim(d)
        chave = sem_inicio.isoformat()
        if chave not in semanas_no_mes:
            semanas_no_mes[chave] = 0.0
        semanas_no_mes[chave] += r['horas_trabalhadas']
    # Adicionar horas justificadas por semana
    d = inicio_mes
    while d <= min(fim_mes, hoje_dt):
        sem_inicio, sem_fim = get_semana_inicio_fim(d)
        chave = sem_inicio.isoformat()
        if chave not in semanas_no_mes:
            semanas_no_mes[chave] = 0.0
        d += timedelta(days=1)
    # Somar justificadas por semana
    for chave in semanas_no_mes:
        sem_inicio = date.fromisoformat(chave)
        sem_fim = sem_inicio + timedelta(days=6)
        hj, _ = calcular_horas_justificadas(user_id, sem_inicio, sem_fim, colaborador, db)
        semanas_no_mes[chave] += hj
    for h in semanas_no_mes.values():
        horas_extras_mes += calcular_horas_extras_semana(h, max_horas)

    # Tipo do dia (normal ou especial)
    tipo_dia_hoje = tipo_dia(hoje_dt, db)
    feriado_hoje = is_feriado(hoje_dt, db)

    # Duração mínima do almoço
    min_almoco = 30 if tipo_dia_hoje == 'especial' else 60

    # Calcular horário mínimo de retorno do almoço (se estiver em almoço)
    retorno_minimo = None
    if registro_hoje and registro_hoje['saida_almoco'] and not registro_hoje['retorno_almoco']:
        fmt = '%H:%M'
        h_saida_almoco = datetime.strptime(registro_hoje['saida_almoco'], fmt)
        h_retorno_min = h_saida_almoco + timedelta(minutes=min_almoco)
        retorno_minimo = h_retorno_min.strftime('%H:%M')

    # Determinar próximo tipo de batida
    proximo_tipo = _determinar_proximo_tipo(registro_hoje)

    # Justificativas pendentes
    justificativas = db.execute(
        '''SELECT * FROM justificativas
           WHERE colaborador_id = ?
           ORDER BY data_registro DESC LIMIT 10''',
        (user_id,)
    ).fetchall()

    db.close()

    return render_template('meu_ponto.html',
                           colaborador=colaborador,
                           registro_hoje=registro_hoje,
                           registros_semana=registros_semana,
                           registros_mes=registros_mes,
                           horas_semana=round(horas_semana, 2),
                           horas_mes=round(horas_mes, 2),
                           horas_just_semana=horas_just_semana,
                           horas_just_mes=horas_just_mes,
                           resumo_sem=resumo_sem,
                           horas_extras_mes=round(horas_extras_mes, 2),
                           tipo_dia_hoje=tipo_dia_hoje,
                           feriado_hoje=feriado_hoje,
                           min_almoco=min_almoco,
                           retorno_minimo=retorno_minimo,
                           proximo_tipo=proximo_tipo,
                           justificativas=justificativas)


def _determinar_proximo_tipo(registro):
    """Determina qual é a próxima batida a ser feita."""
    if not registro:
        return 'entrada'
    if not registro['entrada']:
        return 'entrada'
    if not registro['saida_almoco']:
        return 'saida_almoco'
    if not registro['retorno_almoco']:
        return 'retorno_almoco'
    if not registro['saida']:
        return 'saida'
    return 'completo'


LABELS_TIPO = {
    'entrada': 'Entrada',
    'saida_almoco': 'Saída Almoço',
    'retorno_almoco': 'Retorno Almoço',
    'saida': 'Saída',
    'completo': 'Dia Completo',
}


@app.route('/registrar-ponto', methods=['POST'])
@login_required
def registrar_ponto():
    db = get_db()
    user_id = session['user_id']
    hoje_iso = hoje().isoformat()
    agora_str = agora().strftime('%H:%M')
    td = tipo_dia(hoje(), db)

    # Busca registro de hoje
    registro = db.execute(
        'SELECT * FROM registros_ponto WHERE colaborador_id = ? AND data = ?',
        (user_id, hoje_iso)
    ).fetchone()

    proximo_tipo = _determinar_proximo_tipo(registro)

    if proximo_tipo == 'completo':
        flash('Todas as batidas do dia já foram registradas.', 'info')
        db.close()
        return redirect(url_for('meu_ponto'))

    # Validar duração mínima do almoço ao retornar
    if proximo_tipo == 'retorno_almoco' and registro and registro['saida_almoco']:
        fmt = '%H:%M'
        h_saida_almoco = datetime.strptime(registro['saida_almoco'], fmt)
        h_retorno = datetime.strptime(agora_str, fmt)
        almoco_minutos = (h_retorno - h_saida_almoco).total_seconds() / 60
        min_almoco = 30 if td == 'especial' else 60
        if almoco_minutos < min_almoco:
            restante = int(min_almoco - almoco_minutos)
            flash(f'Intervalo de almoço mínimo: {min_almoco} minutos. '
                  f'Faltam {restante} min. Aguarde para registrar o retorno.', 'warning')
            db.close()
            return redirect(url_for('meu_ponto'))

    if not registro:
        # Criar registro do dia
        db.execute(
            '''INSERT INTO registros_ponto (colaborador_id, data, entrada, tipo_dia, status)
               VALUES (?, ?, ?, ?, 'em_andamento')''',
            (user_id, hoje_iso, agora_str, td)
        )
        flash(f'Entrada registrada às {agora_str}!', 'success')
    else:
        # Atualizar registro existente
        db.execute(
            f'UPDATE registros_ponto SET {proximo_tipo} = ? WHERE id = ?',
            (agora_str, registro['id'])
        )

        # Se é a saída final, calcular horas e marcar como completo
        if proximo_tipo == 'saida':
            reg_atualizado = db.execute(
                'SELECT * FROM registros_ponto WHERE id = ?', (registro['id'],)
            ).fetchone()
            horas = calcular_horas(
                reg_atualizado['entrada'],
                reg_atualizado['saida_almoco'],
                reg_atualizado['retorno_almoco'],
                agora_str
            )
            db.execute(
                '''UPDATE registros_ponto
                   SET horas_trabalhadas = ?, status = 'completo'
                   WHERE id = ?''',
                (horas, registro['id'])
            )
        else:
            # Atualizar horas parciais
            reg_atualizado = db.execute(
                'SELECT * FROM registros_ponto WHERE id = ?', (registro['id'],)
            ).fetchone()
            horas = calcular_horas(
                reg_atualizado['entrada'],
                reg_atualizado['saida_almoco'] if proximo_tipo != 'saida_almoco' else agora_str,
                reg_atualizado['retorno_almoco'] if proximo_tipo != 'retorno_almoco' else agora_str,
                reg_atualizado['saida']
            )
            db.execute(
                'UPDATE registros_ponto SET horas_trabalhadas = ? WHERE id = ?',
                (horas, registro['id'])
            )

        flash(f'{LABELS_TIPO[proximo_tipo]} registrada às {agora_str}!', 'success')

    db.commit()
    db.close()
    return redirect(url_for('meu_ponto'))


# ---------------------------------------------------------------------------
# Manager Dashboard
# ---------------------------------------------------------------------------

@app.route('/dashboard')
@gestor_required
def dashboard():
    db = get_db()
    hoje_iso = hoje().isoformat()
    inicio_sem, fim_sem = get_semana_inicio_fim(hoje())
    inicio_mes, fim_mes = get_mes_inicio_fim(hoje())

    # Todos os colaboradores ativos
    colaboradores = db.execute(
        'SELECT * FROM colaboradores WHERE ativo = 1 ORDER BY nome'
    ).fetchall()

    # Registros de hoje de todos
    registros_hoje = db.execute(
        '''SELECT r.*, c.nome as colaborador_nome
           FROM registros_ponto r
           JOIN colaboradores c ON r.colaborador_id = c.id
           WHERE r.data = ? AND c.ativo = 1
           ORDER BY c.nome''',
        (hoje_iso,)
    ).fetchall()

    # IDs que registraram hoje
    ids_com_registro = {r['colaborador_id'] for r in registros_hoje}

    # Resumo semanal por colaborador (com horas extras)
    resumo_semanal = db.execute(
        '''SELECT c.id, c.nome, c.cargo, c.max_horas_semana, c.folgas_semana,
                  SUM(r.horas_trabalhadas) as horas_semana,
                  COUNT(r.id) as dias_semana
           FROM colaboradores c
           LEFT JOIN registros_ponto r ON c.id = r.colaborador_id
               AND r.data BETWEEN ? AND ?
           WHERE c.ativo = 1
           GROUP BY c.id
           ORDER BY c.nome''',
        (inicio_sem.isoformat(), fim_sem.isoformat())
    ).fetchall()

    # Resumo mensal por colaborador
    resumo_mensal = db.execute(
        '''SELECT c.id, c.nome, c.cargo, c.max_horas_semana,
                  SUM(r.horas_trabalhadas) as horas_mes,
                  COUNT(r.id) as dias_mes
           FROM colaboradores c
           LEFT JOIN registros_ponto r ON c.id = r.colaborador_id
               AND r.data BETWEEN ? AND ?
           WHERE c.ativo = 1
           GROUP BY c.id
           ORDER BY c.nome''',
        (inicio_mes.isoformat(), fim_mes.isoformat())
    ).fetchall()

    # Calcular horas extras mensais e horas justificadas por colaborador
    horas_extras_mensal = {}
    horas_justificadas_mensal = {}
    for c in colaboradores:
        regs_mes = db.execute(
            '''SELECT data, horas_trabalhadas FROM registros_ponto
               WHERE colaborador_id = ? AND data BETWEEN ? AND ?''',
            (c['id'], inicio_mes.isoformat(), fim_mes.isoformat())
        ).fetchall()
        max_h = c['max_horas_semana'] or 40.0

        # Horas justificadas do mês
        hj_mes, _ = calcular_horas_justificadas(c['id'], inicio_mes, fim_mes, c, db)
        horas_justificadas_mensal[c['id']] = hj_mes

        semanas = {}
        for r in regs_mes:
            d = date.fromisoformat(r['data'])
            sem_inicio, _ = get_semana_inicio_fim(d)
            chave = sem_inicio.isoformat()
            semanas[chave] = semanas.get(chave, 0) + r['horas_trabalhadas']
        # Adicionar justificadas por semana
        for chave in list(semanas.keys()):
            sem_inicio = date.fromisoformat(chave)
            sem_fim = sem_inicio + timedelta(days=6)
            hj, _ = calcular_horas_justificadas(c['id'], sem_inicio, sem_fim, c, db)
            semanas[chave] += hj
        extras = sum(calcular_horas_extras_semana(h, max_h) for h in semanas.values())
        horas_extras_mensal[c['id']] = round(extras, 2)

    # Justificativas pendentes
    justificativas_pendentes = db.execute(
        '''SELECT j.*, c.nome as colaborador_nome
           FROM justificativas j
           JOIN colaboradores c ON j.colaborador_id = c.id
           WHERE j.status = 'pendente'
           ORDER BY j.data_registro DESC'''
    ).fetchall()

    # Alertas: colaboradores sem registro hoje (e que não tem justificativa)
    ausentes = []
    for c in colaboradores:
        if c['id'] not in ids_com_registro and not c['is_gestor']:
            just = db.execute(
                '''SELECT id FROM justificativas
                   WHERE colaborador_id = ? AND ? BETWEEN data_inicio AND data_fim
                   AND status = 'aprovado' ''',
                (c['id'], hoje_iso)
            ).fetchone()
            if not just:
                ausentes.append(c)

    db.close()

    return render_template('dashboard.html',
                           colaboradores=colaboradores,
                           registros_hoje=registros_hoje,
                           resumo_semanal=resumo_semanal,
                           resumo_mensal=resumo_mensal,
                           horas_extras_mensal=horas_extras_mensal,
                           horas_justificadas_mensal=horas_justificadas_mensal,
                           justificativas_pendentes=justificativas_pendentes,
                           ausentes=ausentes,
                           inicio_sem=inicio_sem,
                           fim_sem=fim_sem,
                           inicio_mes=inicio_mes,
                           fim_mes=fim_mes)


@app.route('/relatorio-colaborador/<int:colab_id>')
@gestor_required
def relatorio_colaborador(colab_id):
    db = get_db()

    colaborador = db.execute(
        'SELECT * FROM colaboradores WHERE id = ?', (colab_id,)
    ).fetchone()
    if not colaborador:
        flash('Colaborador não encontrado.', 'danger')
        db.close()
        return redirect(url_for('dashboard'))

    # Período: parâmetro ou mês atual
    mes = request.args.get('mes', hoje().strftime('%Y-%m'))
    try:
        ano, m = mes.split('-')
        data_ref = date(int(ano), int(m), 1)
    except (ValueError, TypeError):
        data_ref = hoje().replace(day=1)

    inicio_mes, fim_mes = get_mes_inicio_fim(data_ref)

    registros = db.execute(
        '''SELECT * FROM registros_ponto
           WHERE colaborador_id = ? AND data BETWEEN ? AND ?
           ORDER BY data''',
        (colab_id, inicio_mes.isoformat(), fim_mes.isoformat())
    ).fetchall()

    justificativas = db.execute(
        '''SELECT * FROM justificativas
           WHERE colaborador_id = ?
           AND (data_inicio BETWEEN ? AND ? OR data_fim BETWEEN ? AND ?)
           ORDER BY data_inicio''',
        (colab_id, inicio_mes.isoformat(), fim_mes.isoformat(),
         inicio_mes.isoformat(), fim_mes.isoformat())
    ).fetchall()

    # Agrupar por semana e calcular horas extras por semana
    max_horas = colaborador['max_horas_semana'] or 40.0
    semanas = {}
    for r in registros:
        d = date.fromisoformat(r['data'])
        sem_inicio, sem_fim = get_semana_inicio_fim(d)
        chave = f"{sem_inicio.isoformat()} a {sem_fim.isoformat()}"
        if chave not in semanas:
            semanas[chave] = {'registros': [], 'total_horas': 0, 'horas_extras': 0,
                              'horas_justificadas': 0,
                              'inicio': sem_inicio, 'fim': sem_fim}
        semanas[chave]['registros'].append(r)
        semanas[chave]['total_horas'] += r['horas_trabalhadas']

    # Adicionar horas justificadas por semana
    for chave in semanas:
        hj, _ = calcular_horas_justificadas(
            colab_id, semanas[chave]['inicio'], semanas[chave]['fim'], colaborador, db)
        semanas[chave]['horas_justificadas'] = hj
        semanas[chave]['total_horas'] += hj

    # Calcular horas extras por semana
    total_horas_extras_mes = 0.0
    for chave in semanas:
        h = semanas[chave]['total_horas']
        extras = calcular_horas_extras_semana(h, max_horas)
        semanas[chave]['horas_extras'] = extras
        semanas[chave]['total_horas'] = round(h, 2)
        total_horas_extras_mes += extras

    total_horas_trabalhadas = sum(r['horas_trabalhadas'] for r in registros)
    horas_just_mes, dias_justificados = calcular_horas_justificadas(
        colab_id, inicio_mes, fim_mes, colaborador, db)
    total_horas_mes = total_horas_trabalhadas + horas_just_mes
    total_dias = len(registros)

    # Calcular horas esperadas no mês
    # Considera dias trabalhados: cada dia pode ser normal (8h) ou especial (6h)
    horas_esperadas = 0.0
    d = inicio_mes
    dias_uteis = 0
    while d <= min(fim_mes, hoje()):
        # Conta todos os dias como possíveis dias de trabalho (horário flexível)
        # mas desconta as folgas semanais (2 por semana estimado)
        horas_esperadas += carga_esperada_dia(d, colaborador, db)
        dias_uteis += 1
        d += timedelta(days=1)

    # Descontar folgas: folgas_semana * semanas no período
    folgas = colaborador['folgas_semana'] or 2
    semanas_periodo = max(1, dias_uteis / 7)
    folgas_total = round(folgas * semanas_periodo)
    # Desconta as folgas usando média da carga diária
    media_carga = horas_esperadas / dias_uteis if dias_uteis > 0 else 8.0
    horas_esperadas -= folgas_total * media_carga
    horas_esperadas = max(0, horas_esperadas)

    # Banco de horas (normais, sem contar extras)
    horas_normais_mes = min(total_horas_mes, total_horas_mes - total_horas_extras_mes)
    banco_horas = round(total_horas_mes - horas_esperadas, 2)

    db.close()

    return render_template('relatorio_colaborador.html',
                           colaborador=colaborador,
                           registros=registros,
                           justificativas=justificativas,
                           semanas=semanas,
                           total_horas_mes=round(total_horas_mes, 2),
                           total_horas_extras_mes=round(total_horas_extras_mes, 2),
                           total_dias=total_dias,
                           horas_esperadas=round(horas_esperadas, 2),
                           horas_justificadas_mes=horas_just_mes,
                           dias_justificados=dias_justificados,
                           banco_horas=banco_horas,
                           mes=mes,
                           inicio_mes=inicio_mes,
                           fim_mes=fim_mes)


# ---------------------------------------------------------------------------
# Employee CRUD
# ---------------------------------------------------------------------------

@app.route('/colaboradores')
@gestor_required
def lista_colaboradores():
    db = get_db()
    colaboradores = db.execute(
        'SELECT * FROM colaboradores ORDER BY nome'
    ).fetchall()
    db.close()
    return render_template('colaboradores.html', colaboradores=colaboradores)


@app.route('/colaboradores/novo', methods=['GET', 'POST'])
@gestor_required
def novo_colaborador():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip().lower()
        cargo = request.form.get('cargo', '').strip()
        departamento = request.form.get('departamento', '').strip()
        max_horas_semana = float(request.form.get('max_horas_semana', 40))
        horas_dia_normal = float(request.form.get('horas_dia_normal', 8))
        horas_dia_especial = float(request.form.get('horas_dia_especial', 6))
        folgas_semana = int(request.form.get('folgas_semana', 2))
        is_gestor = 1 if request.form.get('is_gestor') else 0

        if not nome or not email:
            flash('Nome e e-mail são obrigatórios.', 'danger')
            return render_template('colaborador_form.html', colaborador=None)

        db = get_db()
        try:
            db.execute(
                '''INSERT INTO colaboradores
                   (nome, email, senha, cargo, departamento,
                    max_horas_semana, horas_dia_normal, horas_dia_especial,
                    folgas_semana, is_gestor, primeiro_acesso)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (nome, email, '', cargo, departamento,
                 max_horas_semana, horas_dia_normal, horas_dia_especial,
                 folgas_semana, is_gestor, 1)
            )
            db.commit()
            flash(f'Colaborador "{nome}" cadastrado com sucesso! No primeiro login, será solicitada a criação de senha.', 'success')
        except Exception as e:
            flash(f'Erro ao cadastrar: {e}', 'danger')
        finally:
            db.close()

        return redirect(url_for('lista_colaboradores'))

    return render_template('colaborador_form.html', colaborador=None)


@app.route('/colaboradores/<int:colab_id>/editar', methods=['GET', 'POST'])
@gestor_required
def editar_colaborador(colab_id):
    db = get_db()
    colaborador = db.execute(
        'SELECT * FROM colaboradores WHERE id = ?', (colab_id,)
    ).fetchone()

    if not colaborador:
        flash('Colaborador não encontrado.', 'danger')
        db.close()
        return redirect(url_for('lista_colaboradores'))

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip().lower()
        cargo = request.form.get('cargo', '').strip()
        departamento = request.form.get('departamento', '').strip()
        max_horas_semana = float(request.form.get('max_horas_semana', 40))
        horas_dia_normal = float(request.form.get('horas_dia_normal', 8))
        horas_dia_especial = float(request.form.get('horas_dia_especial', 6))
        folgas_semana = int(request.form.get('folgas_semana', 2))
        is_gestor = 1 if request.form.get('is_gestor') else 0
        ativo = 1 if request.form.get('ativo') else 0
        nova_senha = request.form.get('senha', '').strip()
        resetar_acesso = 1 if request.form.get('resetar_acesso') else 0

        try:
            if nova_senha:
                db.execute(
                    '''UPDATE colaboradores
                       SET nome=?, email=?, senha=?, cargo=?, departamento=?,
                           max_horas_semana=?, horas_dia_normal=?, horas_dia_especial=?,
                           folgas_semana=?, is_gestor=?, ativo=?
                       WHERE id=?''',
                    (nome, email, generate_password_hash(nova_senha), cargo, departamento,
                     max_horas_semana, horas_dia_normal, horas_dia_especial,
                     folgas_semana, is_gestor, ativo, colab_id)
                )
            elif resetar_acesso:
                db.execute(
                    '''UPDATE colaboradores
                       SET nome=?, email=?, senha=?, cargo=?, departamento=?,
                           max_horas_semana=?, horas_dia_normal=?, horas_dia_especial=?,
                           folgas_semana=?, is_gestor=?, ativo=?, primeiro_acesso=1
                       WHERE id=?''',
                    (nome, email, '', cargo, departamento,
                     max_horas_semana, horas_dia_normal, horas_dia_especial,
                     folgas_semana, is_gestor, ativo, colab_id)
                )
            else:
                db.execute(
                    '''UPDATE colaboradores
                       SET nome=?, email=?, cargo=?, departamento=?,
                           max_horas_semana=?, horas_dia_normal=?, horas_dia_especial=?,
                           folgas_semana=?, is_gestor=?, ativo=?
                       WHERE id=?''',
                    (nome, email, cargo, departamento,
                     max_horas_semana, horas_dia_normal, horas_dia_especial,
                     folgas_semana, is_gestor, ativo, colab_id)
                )
            db.commit()
            flash(f'Colaborador "{nome}" atualizado!', 'success')
        except Exception as e:
            flash(f'Erro ao atualizar: {e}', 'danger')
        finally:
            db.close()

        return redirect(url_for('lista_colaboradores'))

    db.close()
    return render_template('colaborador_form.html', colaborador=colaborador)


# ---------------------------------------------------------------------------
# Justificativas (Atestados / Faltas)
# ---------------------------------------------------------------------------

@app.route('/justificativas')
@login_required
def lista_justificativas():
    db = get_db()
    if session.get('is_gestor'):
        justificativas = db.execute(
            '''SELECT j.*, c.nome as colaborador_nome
               FROM justificativas j
               JOIN colaboradores c ON j.colaborador_id = c.id
               ORDER BY j.data_registro DESC'''
        ).fetchall()
    else:
        justificativas = db.execute(
            '''SELECT j.*, c.nome as colaborador_nome
               FROM justificativas j
               JOIN colaboradores c ON j.colaborador_id = c.id
               WHERE j.colaborador_id = ?
               ORDER BY j.data_registro DESC''',
            (session['user_id'],)
        ).fetchall()
    db.close()
    return render_template('justificativas.html', justificativas=justificativas)


@app.route('/justificativas/nova', methods=['GET', 'POST'])
@login_required
def nova_justificativa():
    if request.method == 'POST':
        colab_id = session['user_id']

        data_inicio = request.form.get('data_inicio', '')
        data_fim = request.form.get('data_fim', '')
        tipo = request.form.get('tipo', '')
        descricao = request.form.get('descricao', '')

        if not data_inicio or not data_fim or not tipo:
            flash('Preencha todos os campos obrigatórios.', 'danger')
            db = get_db()
            colaboradores = db.execute(
                'SELECT id, nome FROM colaboradores WHERE ativo = 1 ORDER BY nome'
            ).fetchall() if session.get('is_gestor') else []
            db.close()
            return render_template('justificativa_form.html',
                                   justificativa=None, colaboradores=colaboradores)

        # Upload de arquivo (atestado)
        arquivo_nome = ''
        arquivo = request.files.get('arquivo_atestado')
        if arquivo and arquivo.filename and allowed_file(arquivo.filename):
            ext = arquivo.filename.rsplit('.', 1)[1].lower()
            timestamp = agora().strftime('%Y%m%d_%H%M%S')
            arquivo_nome = f"atestado_{colab_id}_{timestamp}.{ext}"
            arquivo.save(os.path.join(app.config['UPLOAD_FOLDER'], arquivo_nome))

        # Calcular dias
        try:
            d_inicio = date.fromisoformat(data_inicio)
            d_fim = date.fromisoformat(data_fim)
            dias = (d_fim - d_inicio).days + 1
        except ValueError:
            dias = 1

        db = get_db()
        try:
            # Se gestor, auto-aprovar
            status = 'aprovado' if session.get('is_gestor') else 'pendente'
            aprovado_por = session['user_id'] if session.get('is_gestor') else None
            data_aprovacao = agora().isoformat() if session.get('is_gestor') else None

            db.execute(
                '''INSERT INTO justificativas
                   (colaborador_id, data_inicio, data_fim, tipo, descricao,
                    arquivo_atestado, dias, status, aprovado_por, data_aprovacao)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (colab_id, data_inicio, data_fim, tipo, descricao,
                 arquivo_nome, dias, status, aprovado_por, data_aprovacao)
            )
            db.commit()
            if status == 'pendente':
                flash('Justificativa enviada para aprovação do gestor!', 'success')
            else:
                flash('Justificativa registrada e aprovada!', 'success')
        except Exception as e:
            flash(f'Erro: {e}', 'danger')
        finally:
            db.close()

        return redirect(url_for('lista_justificativas'))

    db = get_db()
    colaboradores = db.execute(
        'SELECT id, nome FROM colaboradores WHERE ativo = 1 ORDER BY nome'
    ).fetchall() if session.get('is_gestor') else []
    db.close()
    return render_template('justificativa_form.html',
                           justificativa=None, colaboradores=colaboradores)


@app.route('/uploads/atestados/<filename>')
@login_required
def download_atestado(filename):
    """Serve uploaded atestado files."""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/justificativas/<int:just_id>/aprovar', methods=['POST'])
@gestor_required
def aprovar_justificativa(just_id):
    db = get_db()
    acao = request.form.get('acao', 'aprovar')
    status = 'aprovado' if acao == 'aprovar' else 'rejeitado'

    db.execute(
        '''UPDATE justificativas
           SET status = ?, aprovado_por = ?, data_aprovacao = ?
           WHERE id = ?''',
        (status, session['user_id'], agora().isoformat(), just_id)
    )
    db.commit()
    db.close()

    label = 'aprovada' if status == 'aprovado' else 'rejeitada'
    flash(f'Justificativa {label} com sucesso!', 'success')
    return redirect(url_for('lista_justificativas'))


# ---------------------------------------------------------------------------
# Manual Edit (Gestor)
# ---------------------------------------------------------------------------

@app.route('/registro/<int:reg_id>/editar', methods=['GET', 'POST'])
@gestor_required
def editar_registro(reg_id):
    db = get_db()
    registro = db.execute(
        '''SELECT r.*, c.nome as colaborador_nome
           FROM registros_ponto r
           JOIN colaboradores c ON r.colaborador_id = c.id
           WHERE r.id = ?''',
        (reg_id,)
    ).fetchone()

    if not registro:
        flash('Registro não encontrado.', 'danger')
        db.close()
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        entrada = request.form.get('entrada', '').strip() or None
        saida_almoco = request.form.get('saida_almoco', '').strip() or None
        retorno_almoco = request.form.get('retorno_almoco', '').strip() or None
        saida = request.form.get('saida', '').strip() or None
        observacao = request.form.get('observacao', '').strip()

        horas = calcular_horas(entrada, saida_almoco, retorno_almoco, saida)
        status = 'completo' if entrada and saida else 'em_andamento'

        db.execute(
            '''UPDATE registros_ponto
               SET entrada=?, saida_almoco=?, retorno_almoco=?, saida=?,
                   horas_trabalhadas=?, status=?, observacao=?
               WHERE id=?''',
            (entrada, saida_almoco, retorno_almoco, saida,
             horas, status, observacao, reg_id)
        )
        db.commit()
        flash('Registro atualizado!', 'success')
        db.close()
        return redirect(url_for('relatorio_colaborador', colab_id=registro['colaborador_id']))

    db.close()
    return render_template('editar_registro.html', registro=registro)


# ---------------------------------------------------------------------------
# Feriados (Gestor)
# ---------------------------------------------------------------------------

@app.route('/feriados')
@gestor_required
def lista_feriados():
    db = get_db()
    feriados = db.execute(
        'SELECT * FROM feriados ORDER BY data'
    ).fetchall()
    db.close()
    return render_template('feriados.html', feriados=feriados)


@app.route('/feriados/novo', methods=['POST'])
@gestor_required
def novo_feriado():
    data = request.form.get('data', '').strip()
    descricao = request.form.get('descricao', '').strip()
    if not data:
        flash('Data é obrigatória.', 'danger')
        return redirect(url_for('lista_feriados'))
    db = get_db()
    try:
        db.execute(
            'INSERT INTO feriados (data, descricao) VALUES (?, ?)',
            (data, descricao)
        )
        db.commit()
        flash('Feriado adicionado!', 'success')
    except Exception as e:
        flash(f'Erro: {e}', 'danger')
    finally:
        db.close()
    return redirect(url_for('lista_feriados'))


@app.route('/feriados/<int:fer_id>/excluir', methods=['POST'])
@gestor_required
def excluir_feriado(fer_id):
    db = get_db()
    db.execute('DELETE FROM feriados WHERE id = ?', (fer_id,))
    db.commit()
    db.close()
    flash('Feriado removido!', 'success')
    return redirect(url_for('lista_feriados'))


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

@app.route('/exportar/<int:colab_id>')
@gestor_required
def exportar_excel(colab_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    colaborador = db.execute(
        'SELECT * FROM colaboradores WHERE id = ?', (colab_id,)
    ).fetchone()

    mes = request.args.get('mes', hoje().strftime('%Y-%m'))
    try:
        ano, m = mes.split('-')
        data_ref = date(int(ano), int(m), 1)
    except (ValueError, TypeError):
        data_ref = hoje().replace(day=1)

    inicio_mes, fim_mes = get_mes_inicio_fim(data_ref)

    registros = db.execute(
        '''SELECT * FROM registros_ponto
           WHERE colaborador_id = ? AND data BETWEEN ? AND ?
           ORDER BY data''',
        (colab_id, inicio_mes.isoformat(), fim_mes.isoformat())
    ).fetchall()

    # Horas justificadas
    horas_just, _ = calcular_horas_justificadas(colab_id, inicio_mes, fim_mes, colaborador, db)

    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ponto"

    # Header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:G1')
    ws['A1'] = f"Relatório de Ponto - {colaborador['nome']}"
    ws['A1'].font = Font(bold=True, size=14)

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Período: {inicio_mes.strftime('%d/%m/%Y')} a {fim_mes.strftime('%d/%m/%Y')}"

    headers = ['Data', 'Tipo Dia', 'Entrada', 'Saída Almoço', 'Retorno Almoço', 'Saída', 'Horas', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    total_horas = 0
    for i, reg in enumerate(registros, 5):
        d = date.fromisoformat(reg['data'])
        ws.cell(row=i, column=1, value=d.strftime('%d/%m/%Y')).border = thin_border
        td = reg['tipo_dia'] or 'normal'
        ws.cell(row=i, column=2, value='Dom/Feriado' if td == 'especial' else 'Normal').border = thin_border
        ws.cell(row=i, column=3, value=reg['entrada'] or '-').border = thin_border
        ws.cell(row=i, column=4, value=reg['saida_almoco'] or '-').border = thin_border
        ws.cell(row=i, column=5, value=reg['retorno_almoco'] or '-').border = thin_border
        ws.cell(row=i, column=6, value=reg['saida'] or '-').border = thin_border
        ws.cell(row=i, column=7, value=reg['horas_trabalhadas']).border = thin_border
        ws.cell(row=i, column=8, value=reg['status']).border = thin_border
        total_horas += reg['horas_trabalhadas']

    row_total = len(registros) + 5
    ws.cell(row=row_total, column=6, value="TOTAL TRABALHADO:").font = Font(bold=True)
    ws.cell(row=row_total, column=7, value=round(total_horas, 2)).font = Font(bold=True)

    # Horas justificadas
    row_just = row_total + 1
    ws.cell(row=row_just, column=6, value="HORAS JUSTIFICADAS:").font = Font(bold=True, color="0070C0")
    ws.cell(row=row_just, column=7, value=round(horas_just, 2)).font = Font(bold=True, color="0070C0")

    # Total geral
    row_geral = row_just + 1
    ws.cell(row=row_geral, column=6, value="TOTAL GERAL:").font = Font(bold=True)
    ws.cell(row=row_geral, column=7, value=round(total_horas + horas_just, 2)).font = Font(bold=True)

    # Calcular horas extras (incluindo justificadas)
    max_h = colaborador['max_horas_semana'] or 40.0
    semanas_exp = {}
    for reg in registros:
        d = date.fromisoformat(reg['data'])
        sem_inicio, _ = get_semana_inicio_fim(d)
        chave = sem_inicio.isoformat()
        semanas_exp[chave] = semanas_exp.get(chave, 0) + reg['horas_trabalhadas']
    total_extras = sum(calcular_horas_extras_semana(h, max_h) for h in semanas_exp.values())

    row_extras = row_geral + 1
    ws.cell(row=row_extras, column=6, value="HORAS EXTRAS:").font = Font(bold=True, color="FF0000")
    ws.cell(row=row_extras, column=7, value=round(total_extras, 2)).font = Font(bold=True, color="FF0000")

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"ponto_{colaborador['nome'].replace(' ', '_')}_{mes}.xlsx"
    return send_file(output, as_attachment=True, download_name=nome_arquivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------------------------------------------------------------------------
# Alterar Senha (colaborador)
# ---------------------------------------------------------------------------

@app.route('/alterar-senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    if request.method == 'POST':
        senha_atual = request.form.get('senha_atual', '')
        nova_senha = request.form.get('nova_senha', '')
        confirmar = request.form.get('confirmar_senha', '')

        if nova_senha != confirmar:
            flash('As senhas não conferem.', 'danger')
            return render_template('alterar_senha.html')

        if len(nova_senha) < 4:
            flash('A senha deve ter pelo menos 4 caracteres.', 'danger')
            return render_template('alterar_senha.html')

        db = get_db()
        user = db.execute(
            'SELECT * FROM colaboradores WHERE id = ?', (session['user_id'],)
        ).fetchone()

        if not check_password_hash(user['senha'], senha_atual):
            flash('Senha atual incorreta.', 'danger')
            db.close()
            return render_template('alterar_senha.html')

        db.execute(
            'UPDATE colaboradores SET senha = ? WHERE id = ?',
            (generate_password_hash(nova_senha), session['user_id'])
        )
        db.commit()
        db.close()
        flash('Senha alterada com sucesso!', 'success')
        return redirect(url_for('meu_ponto'))

    return render_template('alterar_senha.html')


# ---------------------------------------------------------------------------
# API endpoint para relógio em tempo real
# ---------------------------------------------------------------------------

@app.route('/api/hora-atual')
def hora_atual():
    return jsonify({'hora': agora().strftime('%H:%M:%S')})


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
else:
    # Production (gunicorn) - init DB on import
    init_db()
